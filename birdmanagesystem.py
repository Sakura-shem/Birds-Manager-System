import openpyxl
import os
from tkinter import *
import matplotlib.pyplot as plt
import easygui as g





def draw(mydict01, mydict02, name):

    path = os.path.join(os.path.dirname(__file__), "Image")
    def GetData(mydict):
        ColumnX = []
        ColumnY = []
        for i, j  in zip(mydict.keys(), mydict.values()):
            ColumnX.append(i)
            ColumnY.append(int(j))
        print(ColumnX, ColumnY)
        return ColumnX, ColumnY

    def auto_text(rects):
        for rect in rects:
            ax.text(rect.get_x() + 0.35, rect.get_height()+0.5, rect.get_height())

    # # 柱状图
    ColumnX, ColumnY = GetData(mydict01)
    fig, ax = plt.subplots()
    rect = plt.bar(range(len(ColumnY)), ColumnY, tick_label = ColumnX)
    plt.title("各鸟类数量分析")
    plt.xlabel("鸟类")
    plt.ylabel("数量")
    plt.rcParams["font.sans-serif"]=['SimHei']
    plt.rcParams["axes.unicode_minus"]=False
    auto_text(rect)
    plt.savefig(path + "\\Rect-" + name)
    

    # 折线图
    plt.rcParams["font.sans-serif"]=['SimHei']
    plt.rcParams["axes.unicode_minus"]=False
    ColumnX, ColumnY = GetData(mydict02)
    plt.figure(figsize=(20,5),dpi=90)
    plt.plot(ColumnX, ColumnY)
    plt.xticks(rotation=45)
    plt.title("活动量分析")
    plt.xlabel("日期")
    plt.ylabel("鸟类活动总数")
    plt.savefig(path + "\\line-" + name)
    

def openfile(x):
    #打开文件名为x(不含路径)的文件
    if x[-4:]=='xlsx':
        path = os.path.join(os.path.dirname(__file__), "Data")
        cmd = 'start ' +path + '\\'+x
        os.system(cmd)
    if x[-4:]=='.png':
        path = os.path.join(os.path.dirname(__file__), "Image")
        cmd = 'start ' +path + '\\'+x
        os.system(cmd)
def findbirdsdata(e1,t):
    t.insert(INSERT,'此鸟类数据如下:\n')
    path = os.path.join(os.path.dirname(__file__), "Data")
    for eachfile in os.listdir(path):
        if os.path.splitext(eachfile)[1]=='.xlsx':
            print(eachfile)
            wb=openpyxl.load_workbook(os.path.join(path, eachfile))
            ws=wb.active
            for row in ws.rows:
                if row[1].value==e1:
                    t.insert(INSERT,'日期:%s  鸟种:%s  观测点:%s  数量:%s  飞行高度:%s  活动状态:%s  活动环境:%s  驱赶方式:%s  物资消耗:%s \n'%(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,row[8].value))

def localdata():
    root=Tk()
    root.geometry('400x300')
    root.title('本地数据')
    l=Listbox(root,width=40)
    l.pack()
    path0 = os.path.join(os.path.dirname(__file__), "Data")
    for eachfile in os.listdir(path0):
        file=os.path.basename(eachfile)
        if file[-5:-1:1] == ".xls":
            l.insert(END,file)
    path1 = os.path.join(os.path.dirname(__file__), "Image")
    for eachfile in os.listdir(path1):
        file=os.path.basename(eachfile)
        if file[-4:] == ".png":
            l.insert(END,file)
    Button(root,text='打开',command=lambda:openfile(l.get(ACTIVE))).pack()

def searchdata():
    root=Tk()
    root.geometry('900x300')
    root.title('搜索鸟类数据')
    Label(root,text='请输入鸟的种类:').grid(row=0,column=0,sticky=W)
    e1=Entry(root)
    e1.grid(row=0,column=1, sticky = W)
    t=Text(root,width=130, height = 20)
    t.grid(row = 2, column = 0, columnspan=3)
    Button(root,text='开始查询',command=lambda:findbirdsdata(e1.get(),t)).grid(row=1,sticky=W)
def savethisdata(EntryList):
    EntryText = [x.get().strip() for x in EntryList]
    if '' in EntryText:
        print("1")
        return
    filename = EntryText[0][0:4] + '年' + EntryText[0][5:7] + '月鸟类数据.xlsx'
    photoname = EntryText[0][0:4] + '年' + EntryText[0][5:7] + '月鸟类图解.png'
    birdnames = {}
    dataname = {}
    global path
    path0 = os.path.join(path, "Data")
    newpath = os.path.join(path, "Data", filename)
    if filename not in os.listdir(path0):
        wb=openpyxl.Workbook()
        ws=wb.active
        ws['A1']='日期'
        ws['B1']='鸟种'
        ws['C1']='观测点'
        ws['D1']='数量'
        ws['E1']='飞行高度'
        ws['F1']='活动状态'
        ws['G1']='活动环境'
        ws['H1']='驱赶方式'
        ws['I1']='物资消耗'
        ws.append(EntryText)
        for k in ['A','B','C','D','E','F','G','H','I']:
            ws.column_dimensions[k].width = 15
        wb.save(os.path.join(path, "Data", filename))
    else:
        wb=openpyxl.load_workbook(newpath)
        ws=wb.active
        ws.append(EntryText)
        wb.save(os.path.join(path0, filename))
    wc=openpyxl.load_workbook(newpath)
    wd=wc.active
    for row in wd.iter_rows(min_row=2):
        if row[1].value not in birdnames.keys():
            if row[3].value != None:
                birdnames[row[1].value] =int( row[3].value)
        else:
            birdnames[row[1].value]+=int(row[3].value)
        if row[0].value[8:] not in dataname.keys():
            if row[3].value != None:
                dataname[row[0].value[8:]] =int(row[3].value)
        else:
            dataname[row[0].value[8:]] +=int(row[3].value)
    draw(birdnames, dataname, photoname)
    g.msgbox('保存成功！','提示')


def cleardata(Entrylist):
    for i in Entrylist:
        i.delete(0, END)

def inputdata():
    root=Tk()
    root.geometry('300x300')
    root.title('请输入本月数据')
    EntryList = [Entry(root, width = 34) for x in range(0, 9)]
    TextList = ["日期", "鸟种", "观测点", "数量", "飞行高度", "活动状态", "活动环境", "驱赶方式", "物资消耗"]
    LabelList = [Label(root, text = i).grid(row = TextList.index(i)) for i in TextList]
    for i in EntryList:
        i.grid(row = EntryList.index(i), column = 1)
    Button(root,text='保存到本月表单',command=lambda:savethisdata(EntryList)).grid(row=9,column=0)
    Button(root,text='清空本次数据   ',command=lambda:cleardata(EntryList)).grid(row=10,sticky=W)

if __name__ == "__main__":
    path = os.path.dirname(__file__)
    root=Tk()
    root.title('鸟类资源管理器')
    root.geometry('400x90')
    Button(root,text='输入数据',bg='white',command=inputdata).pack(fill=X)
    Button(root,text='本地数据',bg='white',command=localdata).pack(fill=X)
    Button(root,text='搜索',bg='white',command=searchdata).pack(fill=X)
    mainloop()
